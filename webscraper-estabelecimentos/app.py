"""
Aplicacao web (Flask) para o scraper do Google Maps.

Coleta Nome, Telefone, Cidade - UF e CEP dos estabelecimentos,
mostrando APENAS os que NAO possuem website.

Uso:
    python app.py
    -> abra http://127.0.0.1:5000
"""

import io
import json

from flask import (Flask, Response, jsonify, render_template, request,
                   send_file, stream_with_context)

from scraper import scrape_batch

app = Flask(__name__)


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/scrape")
def scrape_stream():
    # Aceita varias buscas: ?q=...&q=...  (modo lote)
    queries = [q.strip() for q in request.args.getlist("q") if q and q.strip()]
    if not queries:  # compatibilidade com o parametro antigo
        unica = (request.args.get("query") or "").strip()
        if unica:
            queries = [unica]
    try:
        max_results = int(request.args.get("max", 40))
    except ValueError:
        max_results = 40
    max_results = max(1, min(max_results, 200))
    headless = request.args.get("headless", "1") != "0"
    exigir_telefone = request.args.get("telefone", "1") != "0"

    def gerar():
        if not queries:
            yield _sse("error", "Informe ao menos um termo de busca.")
            yield _sse("done", {"processados": 0, "encontrados": 0, "buscas": 0})
            return
        try:
            for tipo, dado in scrape_batch(queries, max_results=max_results,
                                           headless=headless,
                                           exigir_telefone=exigir_telefone):
                yield _sse(tipo, dado)
        except Exception as e:  # erro fatal (ex.: navegador nao instalado)
            yield _sse("error", f"Erro fatal: {type(e).__name__}: {e}")
            yield _sse("done", {"processados": 0, "encontrados": 0, "buscas": 0})

    return Response(stream_with_context(gerar()),
                    mimetype="text/event-stream",
                    headers={"Cache-Control": "no-cache",
                             "X-Accel-Buffering": "no"})


def _sse(tipo, dado):
    return f"event: {tipo}\ndata: {json.dumps(dado, ensure_ascii=False)}\n\n"


@app.route("/export/xlsx", methods=["POST"])
def export_xlsx():
    """Recebe a lista de resultados (JSON) e devolve um arquivo .xlsx."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    dados = request.get_json(silent=True) or []
    if not isinstance(dados, list):
        return jsonify({"erro": "Formato inválido."}), 400

    wb = Workbook()
    ws = wb.active
    ws.title = "Sem site"

    colunas = [("Nome", 38), ("Telefone", 18), ("WhatsApp", 30),
               ("Cidade - UF", 24), ("CEP", 12), ("Endereço", 55)]
    chaves = ["nome", "telefone", "whatsapp", "cidade_uf", "cep", "endereco"]

    cabecalho_fill = PatternFill("solid", fgColor="1E293B")
    cabecalho_font = Font(bold=True, color="FFFFFF")
    link_font = Font(color="0563C1", underline="single")
    for i, (titulo, largura) in enumerate(colunas, start=1):
        c = ws.cell(row=1, column=i, value=titulo)
        c.fill = cabecalho_fill
        c.font = cabecalho_font
        c.alignment = Alignment(vertical="center")
        ws.column_dimensions[get_column_letter(i)].width = largura

    for linha, item in enumerate(dados, start=2):
        if not isinstance(item, dict):
            continue
        for col, chave in enumerate(chaves, start=1):
            valor = str(item.get(chave, ""))
            cel = ws.cell(row=linha, column=col, value=valor)
            if chave == "whatsapp" and valor.startswith("http"):
                cel.hyperlink = valor
                cel.value = "Abrir WhatsApp"
                cel.font = link_font

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(colunas))}{max(1, len(dados) + 1)}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="estabelecimentos_sem_site.xlsx",
    )


if __name__ == "__main__":
    app.run(debug=True, threaded=True, port=5000)
